from openpyxl import Workbook
import random

def create_advanced_dataset():
    wb = Workbook()

    # --- Sheet 1: Clients (Main Data) ---
    ws_clients = wb.active
    ws_clients.title = "Clients_Brut"

    headers = ["ID_Client", "Nom_Complet", "Date_Souscription", "Code_Offre", "Region", "Statut_Paiement"]
    ws_clients.append(headers)

    regions = ["IDF", "PACA", "BRETAGNE", "NORD", "EST", "SUD-OUEST"]
    offres = ["FIBRE_1GB", "ADSL_20M", "MOBILE_50GB", "MOBILE_100GB", "FIBRE_PRO"]
    statuts = ["OK", "Retard", "Contentieux", "OK", "OK"]

    # Generate 50 rows of dirty data
    for i in range(1, 51):
        client_id = f"C{i:04d}"
        if i % 10 == 0: client_id = f"C{i:04d}" # Duplicate ID scenario

        prenom = f"Prenom{i}"
        nom = f"NOM{i}"
        nom_complet = f"{nom} {prenom}" if i % 3 != 0 else f"{prenom} {nom}" # Inconsistent name order

        date = f"2023-{random.randint(1,12):02d}-{random.randint(1,28):02d}"
        if i % 5 == 0: date = f"{random.randint(1,28):02d}/{random.randint(1,12):02d}/2023" # Mixed Date format

        offre = random.choice(offres)
        region = random.choice(regions) if i % 8 != 0 else "" # Missing region
        statut = random.choice(statuts)

        ws_clients.append([client_id, nom_complet, date, offre, region, statut])

    # --- Sheet 2: Offres (Lookup Table) ---
    ws_offres = wb.create_sheet("Ref_Offres")
    ws_offres.append(["Code_Offre", "Prix_Mensuel", "Engagement_Mois", "Type"])

    data_offres = [
        ["FIBRE_1GB", 39.99, 12, "Fixe"],
        ["ADSL_20M", 29.99, 12, "Fixe"],
        ["MOBILE_50GB", 19.99, 0, "Mobile"],
        ["MOBILE_100GB", 25.99, 24, "Mobile"],
        ["FIBRE_PRO", 89.00, 24, "Pro"]
    ]

    for row in data_offres:
        ws_offres.append(row)

    # --- Sheet 3: Logs Appels (For Pivot Table & Cross-Ref) ---
    ws_logs = wb.create_sheet("Logs_Appels")
    ws_logs.append(["ID_Appel", "ID_Client", "Duree_Minutes", "Date_Appel"])

    for i in range(1, 101):
        client_ref = f"C{random.randint(1,50):04d}"
        ws_logs.append([f"CALL{i:05d}", client_ref, random.randint(1, 120), "2024-01-15"])

    filename = "dataset_clients_advanced.xlsx"
    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    create_advanced_dataset()
