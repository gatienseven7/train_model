from openpyxl import Workbook
import random

def create_week4_dataset():
    wb = Workbook()

    # --- Sheet 1: Notes Elèves (Pour la fonction SI) ---
    ws_notes = wb.active
    ws_notes.title = "Notes_Examens"
    ws_notes.append(["Nom_Eleve", "Note_Maths", "Note_Francais", "Moyenne", "Resultat_Attendu"])

    noms = ["Alice", "Bob", "Charlie", "David", "Emma", "Fabien", "Gael", "Hannah", "Ines", "Jules"]
    for nom in noms:
        maths = random.randint(5, 20)
        francais = random.randint(5, 20)
        moyenne = (maths + francais) / 2
        resultat = "Admis" if moyenne >= 10 else "Recalé"
        ws_notes.append([nom, maths, francais, moyenne, resultat])

    # --- Sheet 2: Ventes Magasin (Pour les Graphiques Histogramme) ---
    ws_ventes = wb.create_sheet("Ventes_Mensuelles")
    ws_ventes.append(["Vendeur", "Janvier", "Fevrier", "Mars", "Total_Trimestre"])

    vendeurs = ["Pierre", "Paul", "Jacques", "Marie", "Sophie"]
    for v in vendeurs:
        jan = random.randint(1000, 5000)
        feb = random.randint(1000, 5000)
        mar = random.randint(1000, 5000)
        total = jan + feb + mar
        ws_ventes.append([v, jan, feb, mar, total])

    # --- Sheet 3: Budget Depenses (Pour le Camembert - Rappel S2) ---
    ws_budget = wb.create_sheet("Budget_Projet")
    ws_budget.append(["Categorie", "Depense_Prevue", "Depense_Reelle"])
    categories = ["Loyer", "Salaires", "Marketing", "Logiciels", "Divers"]
    for cat in categories:
        prevu = random.randint(2000, 10000)
        reel = prevu + random.randint(-500, 1000)
        ws_budget.append([cat, prevu, reel])

    filename = "dataset_S4_logique_viz.xlsx"
    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    create_week4_dataset()
